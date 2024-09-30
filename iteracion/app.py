import openpyxl
from itertools import product

# Cargar el archivo Excel
workbook = openpyxl.load_workbook("WebScraping_Finding/src/Libro2.xlsx")

# Definir configuraciones de columnas para cada hoja 
configuraciones = { 
    "Gonzalo Mejía": {"col1_range": range(1), "col2_range": range(2)}, 
    "Alejandro Acosta": {"col1_range": range(1), "col2_range": range(2)}, 
    "Victor García": {"col1_range": range(1), "col2_range": range(2)}, 
    "Ximena Quintanilla": {"col1_range": range(1), "col2_range": range(2)}, 
    "Éder Caceres": {"col1_range": range(1), "col2_range": range(2)}, 
    "Diego Jaimes": {"col1_range": range(1), "col2_range": range(2)}, 
    "Felipe Reyes": {"col1_range": range(1), "col2_range": range(2)}, 
}

# Crear un nuevo libro de Excel para almacenar las combinaciones
nuevo_libro = openpyxl.Workbook()
nueva_hoja = nuevo_libro.active
nueva_hoja.title = "Combinaciones"

# Inicializar la fila para empezar a escribir en la nueva hoja
fila_actual = 1

# Iterar sobre cada hoja del archivo Excel
for hoja in workbook.worksheets:
    data = []

    # Leer los datos desde la hoja actual (suponiendo que los datos comienzan en la primera fila)
    for row in range(0, hoja.max_row):
        _row = []
        for col in hoja.iter_cols(1, hoja.max_column + 1):
            _row.append(col[row].value)
        data.append(_row)

    num_filas = len(data)

    # Obtener la configuración específica para la hoja actual
    config = configuraciones.get(hoja.title)

    if config and num_filas > 0:
        # Escribir el nombre de la hoja como un encabezado en la nueva hoja de Excel
        #nueva_hoja.cell(row=fila_actual, column=1, value=f"--- Combinaciones para la hoja: {hoja.title} ---")
        #fila_actual += 1  # Avanzar a la siguiente fila

        # Generar combinaciones entre dos columnas diferentes
        for col1 in config["col1_range"]:
            for col2 in config["col2_range"]:
                # Solo generar combinaciones entre columnas diferentes
                if col1 != col2:  # Esta condición asegura que las columnas sean distintas
                    col1_values = [data[row][col1] for row in range(num_filas) if len(data[row]) > col1 and data[row][col1] is not None]
                    col2_values = [data[row][col2] for row in range(num_filas) if len(data[row]) > col2 and data[row][col2] is not None]

                    # Generar y escribir combinaciones de dos columnas
                    for val1, val2 in product(col1_values, col2_values):
                        nueva_hoja.cell(row=fila_actual, column=1, value=f'"{val1}"  "{val2}"')
                        fila_actual += 1  # Avanzar a la siguiente fila

# Guardar el nuevo archivo de Excel con las combinaciones
nuevo_libro.save("WebScraping_Finding/src/Combinaciones_Generadas.xlsx")
print("Las combinaciones se han guardado en el archivo 'Combinaciones_Generadas.xlsx'.")
